#!/usr/bin/env python

import glob
import os
import re
import json
import openpyxl
import openpyxl
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side


class ReportHelper:
    """
    @author: RexGuo
    @description: a module to generate fio and iostat report into an Excel file.
    """

    FIO_TEMPLATE_HEADER_ROW = 1
    FIO_TEMPLATE_DATA_START_ROW = 2
    IOSTAT_TEMPLATE_HEADER_ROW = 1  # 表头所在行(测试1..)
    IOSTAT_TEMPLATE_INDEX_ROW = IOSTAT_TEMPLATE_HEADER_ROW + 1  # 指标所在行(IOPS..)
    IOSTAT_TEMPLATE_DATA_START_ROW = 3  # 数据起始行（图表起始行）
    HEADER_COLOR = "92D050"
    FIO_SHEET_NAME = "FIO_Latency"

    def __init__(self, **kwargs):
        """
        @param: log_path : log文件所在目录
        @param: iostat_data : iostat数据
        @param: rpt_path : 报告文件输出路径
        """
        self.log_path = kwargs.get("log_path")
        self.iostat_data = kwargs.get("iostat_data")
        self.rpt_path = kwargs.get("rpt_path")
        self.wb = openpyxl.Workbook()

    def generate_fio_template(self, disk_data):
        for k, v in disk_data.items():
            diskname = k
        sht = self.wb.create_sheet(title=f"FIO_{diskname}")
        headers_1 = [
            "测试顺序",
            "测试模型说明",
            "",
            "Min Lat\n（μs）",
            "Max Lat\n（μs）",
            "Ave. Lat\n（μs）",
            "STDEV Lat\n（μs）",
            "CPU-user",
            "CPU-sys",
            "产品手册值",
            "稳态值取值范围",
            "",
            "稳态平均值1",
            "稳态平均值2",
            "比例",
            "均值结果",
            "抖动点比例",
            "抖动结果",
            "稳态内\n最小值",
            "稳态内\n最大值",
            "最大跌落\n点结果",
            "趋势结果",
        ]
        # headers_2 = [
        #     "",
        #     "",
        #     "",
        #     "（μs）",
        #     "（μs）",
        #     "（μs）",
        #     "（μs）",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "最小值",
        #     "最大值",
        #     "点结果",
        #     "",
        # ]
        # 填入表头
        yellow_fill_col = ["均值结果", "抖动结果", "最大跌落\n点结果", "趋势结果"]
        for i, v in enumerate(headers_1):
            sht.cell(1, i + 1).value = v
            sht.cell(1, i + 1).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
                shrink_to_fit=True,
            )
            if v in yellow_fill_col:
                sht.cell(1, i + 1).fill = PatternFill(
                    start_color="ffff00",
                    end_color="ffff00",
                    fill_type="solid",
                )
                continue
            sht.cell(1, i + 1).fill = PatternFill(
                start_color=self.HEADER_COLOR,
                end_color=self.HEADER_COLOR,
                fill_type="solid",
            )

        # for i, v in enumerate(headers_2):
        #     sht.cell(2, i + 1).value = v
        #     sht.cell(2, i + 1).alignment = Alignment(
        #         horizontal="center",
        #         vertical="center",
        #         wrap_text=True,
        #         shrink_to_fit=True,
        #     )

        # 合并单元格
        # 第二行没值的: 直接合并上下两行
        # 遍历单元格，分别看两行，如果两行都为空则和前方列合并
        for i in range(1, len(headers_1) + 1):
            header_1 = sht.cell(1, i).value
            if header_1 == "":
                sht.merge_cells(
                    start_column=i - 1, start_row=1, end_column=i, end_row=1
                )
        # for i in range(1, len(headers_1) + 1):
        #     header_1 = sht.cell(1, i).value
        #     header_2 = sht.cell(2, i).value
        #     if header_2 == "":
        #         if header_1 == "":
        #             sht.merge_cells(
        #                 start_column=i - 1, start_row=1, end_column=i, end_row=2
        #             )
        #             # sht.merge_cells(
        #             #     "{h1}1:{h2}2".format(
        #             #         h1=get_column_letter(i - 1), h2=get_column_letter(i)
        #             #     )
        #             # )
        #         else:
        #             column_letter = get_column_letter(i)
        #             sht.merge_cells(
        #                 start_column=i, start_row=1, end_column=i, end_row=2
        #             )
        #             # sht.merge_cells(f"{column_letter}1:{column_letter}2")

    def generate_iostat_report(self, data):
        # sht = self.wb.active
        ######FOR COMPATIBILITY WITH OTHER FORMAT
        testdata = []
        devs = []
        for dev in data:
            for k, v in dev.items():
                if k not in devs:
                    devs.append(k)
                    td = {}
                    td[k] = v
                    testdata.append(td)
                else:
                    testdata[devs.index(k)][k].append(v[0])
        data = testdata
        #############################
        for sheet in data:
            current_col = 0
            for k, v in sheet.items():
                iostat_sheet_name = k
                sheet_content = v
            # if iostat_sheet_name != sht.title:
            sht = self.wb.create_sheet(title=f"IOSTAT_{iostat_sheet_name}")
            for case in sheet_content:
                # 写第一行表头(测试1,测试2..)
                for k, v in case.items():
                    case_name = k
                    case_content = v
                    index_count = len(v)
                start_col = current_col + 1
                end_col = current_col + index_count
                current_col = end_col
                sht.cell(self.IOSTAT_TEMPLATE_HEADER_ROW, start_col).value = case_name
                sht.cell(
                    self.IOSTAT_TEMPLATE_HEADER_ROW, start_col
                ).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                    shrink_to_fit=True,
                )
                sht.cell(self.IOSTAT_TEMPLATE_HEADER_ROW, start_col).fill = PatternFill(
                    start_color=self.HEADER_COLOR,
                    end_color=self.HEADER_COLOR,
                    fill_type="solid",
                )
                sht.merge_cells(
                    start_column=start_col, start_row=1, end_column=end_col, end_row=1
                )
                # 写第二行表头(IOPS,BW..)以及数据
                # 获取表头行号
                for i in range(0, index_count):
                    index_col = start_col + i
                    for k, v in case_content[i].items():
                        index_name = k
                        index_data = v
                    sht.cell(
                        self.IOSTAT_TEMPLATE_INDEX_ROW, index_col
                    ).value = index_name
                    sht.cell(
                        self.IOSTAT_TEMPLATE_INDEX_ROW, index_col
                    ).alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                        shrink_to_fit=True,
                    )
                    for row in range(0, len(index_data)):
                        sht.cell(
                            row + self.IOSTAT_TEMPLATE_DATA_START_ROW, index_col
                        ).value = float(index_data[row])

                # 为所有单元格添加边框
                self._add_borders(sht)

    def _add_borders(self, sht):
        # 为所有单元格添加边框
        for row in range(1, sht.max_row + 1):
            for col in range(1, sht.max_column + 1):
                sht.cell(row, col).border = Border(
                    top=Side(border_style="medium", color="FF000000"),
                    bottom=Side(border_style="medium", color="FF000000"),
                    left=Side(border_style="medium", color="FF000000"),
                    right=Side(border_style="medium", color="FF000000"),
                )

    def add_iostat_chart(self):
        for iostat_sheet_name in self.wb.sheetnames:
            if "IOSTAT" not in iostat_sheet_name:
                continue
            sht = self.wb[iostat_sheet_name]
            col_counts = sht.max_column
            row_counts = sht.max_row
            headers = {}
            for col in range(1, col_counts + 1):
                header = sht.cell(self.IOSTAT_TEMPLATE_HEADER_ROW, col).value
                if header and not header in headers:
                    headers[header] = [col]
                    present_header = header
                if header == None:
                    headers[present_header].append(col)
            chart_start_row = self.IOSTAT_TEMPLATE_DATA_START_ROW
            for header in headers:
                for index_col in headers[header]:
                    index_name = sht.cell(
                        self.IOSTAT_TEMPLATE_INDEX_ROW, index_col
                    ).value
                    # index_data = [ sht.cell(i,index_col).value for i in range(self.IOSTAT_TEMPLATE_DATA_START_ROW,row_counts+1) ]
                    graph_title = f"{header}-{index_name}"
                    chart = LineChart()
                    chart.title = graph_title
                    chart.style = 6 + headers[header].index(index_col)
                    chart.y_axis.title = index_name
                    chart.x_axis.title = None
                    chart.x_axis.delete = True
                    chart.legend = None
                    data = Reference(
                        sht,
                        min_col=index_col,
                        min_row=self.IOSTAT_TEMPLATE_DATA_START_ROW,
                        max_col=index_col,
                        max_row=row_counts,
                    )
                    chart.add_data(data, titles_from_data=False)
                    sht.add_chart(
                        chart, get_column_letter(col_counts + 1) + str(chart_start_row)
                    )
                    chart_start_row += 15

    def fill_fio_data(self, disk_data):
        for k, v in disk_data.items():
            diskname = k
            rows = v
        sht = self.wb[f"FIO_{diskname}"]
        col_counts = sht.max_column
        # 记录表头对应的列
        headers = {}
        for col in range(1, col_counts + 1):
            headers[sht.cell(self.FIO_TEMPLATE_HEADER_ROW, col).value] = col
        # 写数据
        for i in range(0, len(rows)):
            start_row_index = self.FIO_TEMPLATE_DATA_START_ROW + i
            test = rows[i]
            test["测试顺序"] = "测试%d" % (i + 1)
            for k, v in test.items():
                sht.cell(start_row_index, headers[k]).value = v
        self._add_borders(sht)

    def parse_fio_logs(self):
        # 添加单个盘的log路径
        ### 第一层
        disk_dirs = list(
            filter(lambda f: os.path.isdir(f), glob.glob(self.log_path + "/*"))
        )
        all_disk_rows = []
        for dir in disk_dirs:
            ### 第二层
            disk = {}
            fio_sheet_name = os.path.basename(dir)
            logfiles_list = glob.glob(dir + "/*.log")
            unused_file = re.compile(".+\.\d+\.log")
            # 剔除不需要解析的log
            logfiles_list = list(
                filter(lambda f: unused_file.search(f) == None, logfiles_list)
            )
            # log按时间排序处理
            logfiles_list.sort(key=lambda f: os.stat(f).st_mtime)
            rows = []  # 匹配字段列表(行)
            # 解析log文件
            for logfile in logfiles_list:
                test = {}  # 当前测试
                try:
                    with open(logfile, "r") as f:
                        content = json.load(f)["jobs"][0]
                except:
                    continue
                description = os.path.basename(content["jobname"])
                rw = content["job options"]["rw"]
                # 判断 rw 来确定在哪里取值
                if "rw" in rw:
                    description += "_r_w"
                    min_lat = (
                        str(content["read"]["lat_ns"]["min"] / 1000)
                        + ";"
                        + str(content["write"]["lat_ns"]["min"] / 1000)
                    )
                    max_lat = (
                        str(content["read"]["lat_ns"]["max"] / 1000)
                        + ";"
                        + str(content["write"]["lat_ns"]["max"] / 1000)
                    )
                    ave_lat = (
                        str(content["read"]["lat_ns"]["mean"] / 1000)
                        + ";"
                        + str(content["write"]["lat_ns"]["mean"] / 1000)
                    )
                    stddev_lat = (
                        str(content["read"]["lat_ns"]["stddev"] / 1000)
                        + ";"
                        + str(content["write"]["lat_ns"]["stddev"] / 1000)
                    )
                elif "read" in rw:
                    min_lat = content["read"]["lat_ns"]["min"] / 1000
                    max_lat = content["read"]["lat_ns"]["max"] / 1000
                    ave_lat = content["read"]["lat_ns"]["mean"] / 1000
                    stddev_lat = content["read"]["lat_ns"]["stddev"] / 1000
                elif "write" in rw:
                    min_lat = content["write"]["lat_ns"]["min"] / 1000
                    max_lat = content["write"]["lat_ns"]["max"] / 1000
                    ave_lat = content["write"]["lat_ns"]["mean"] / 1000
                    stddev_lat = content["write"]["lat_ns"]["stddev"] / 1000
                cpu_user = content["usr_cpu"]
                cpu_sys = content["sys_cpu"]
                test["测试模型说明"] = description
                test["Min Lat\n（μs）"] = min_lat
                test["Max Lat\n（μs）"] = max_lat
                test["Ave. Lat\n（μs）"] = ave_lat
                test["STDEV Lat\n（μs）"] = stddev_lat
                test["CPU-user"] = cpu_user
                test["CPU-sys"] = cpu_sys
                rows.append(test)
            disk[fio_sheet_name] = rows
            all_disk_rows.append(disk)

        return all_disk_rows

    def generate_fio_report(self):
        all_disk_rows = self.parse_fio_logs()
        for disk in all_disk_rows:
            self.generate_fio_template(disk)
            self.fill_fio_data(disk)

    def generate_report(self):
        self.generate_fio_report()
        self.generate_iostat_report(self.iostat_data)
        self.add_iostat_chart()
        del self.wb[self.wb.sheetnames[0]]
        self.wb.save(self.rpt_path)
        self.wb.close()


# if __name__ == "__main__":
#     import random

#     testData = [
#         {
#             "Sheet1": [
#                 {
#                     "测试1": [
#                         {"IOPS": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试2": [
#                         {"IOPS": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试3": [
#                         {"IOPS_R": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"IOPS_W": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {
#                             "IOPS_Total": [
#                                 random.randint(1, 1000) for i in range(0, 100)
#                             ]
#                         },
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试4": [
#                         {"IOPS": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试5": [
#                         {"IOPS": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试6": [
#                         {"BW_R": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW_W": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW_Total": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {
#                             "IOPS_Total": [
#                                 random.randint(1, 1000) for i in range(0, 100)
#                             ]
#                         },
#                     ]
#                 },
#             ]
#         },
#         {
#             "Sheet2": [
#                 {
#                     "测试1": [
#                         {"IOPS": [random.randint(1, 1000) for i in range(0, 1000)]},
#                         {"BW": [random.randint(1, 1000) for i in range(0, 2000)]},
#                     ]
#                 },
#                 {
#                     "测试2": [
#                         {"IOPS": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试3": [
#                         {"IOPS_R": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"IOPS_W": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {
#                             "IOPS_Total": [
#                                 random.randint(1, 1000) for i in range(0, 100)
#                             ]
#                         },
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试4": [
#                         {"IOPS": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试5": [
#                         {"IOPS": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW": [random.randint(1, 1000) for i in range(0, 100)]},
#                     ]
#                 },
#                 {
#                     "测试6": [
#                         {"BW_R": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW_W": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {"BW_Total": [random.randint(1, 1000) for i in range(0, 100)]},
#                         {
#                             "IOPS_Total": [
#                                 random.randint(1, 1000) for i in range(0, 100)
#                             ]
#                         },
#                     ]
#                 },
#             ]
#         },
#     ]

#     test = ReportHelper(
#         log_path="./logpath", rpt_path="./rpt3.xlsx", iostat_data=testData
#     )
#     test.generate_report()
#     # test.parse_fio_logs()
