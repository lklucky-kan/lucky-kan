#/usr/bin/python3
# -*- coding=utf-8 -*-

import re, os, sys, common, shutil
from openpyxl import load_workbook, Workbook, styles
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
from collections import OrderedDict
from common.other.log import Logger


class Excel():
    '''
    description: this class contain the operation read/write/modify excel with xlsx format
    author: Kail
    '''
    def __init__(self, **kwargs):
        '''
        description: define global variables in class
        author: Kail
        params: kwargs, optional keywords params:
                    logger, object of logger, default log file is csv.log
                    xls_nm, the file path of excel
        '''
        log_file = kwargs.get('log_file', 'excel.log')
        #self.logger = Logger(log_file=log_file)
        if kwargs.get('logger'):
            self.logger = kwargs.get('logger')
        # else:
            # self.logger = Logger(log_file='excel.log')
        self.xls_nm = kwargs.get('xls_nm', 'Sample.xlsx')


    def list_2sheet(self, list_data, sheet_name, **kwargs):
        '''
        description: write pure list data structure in 'xls_nm' Excel's 'sheet_name' Sheet
        author: Kail
        params: list_data, sheet_name
                list_data: pure list data structure
                    eg: [[A1, B1, C1, ...], [A2, B2, C2, ...], [A3, B3, C3, ...]]
                sheet_name: the Sheet name which want to write in
                kwargs: optional keyword value pairs
                    mode: the write mode in sheet, choose from [w,a]
                         w: write mode, cover old data in 'sheet_name' Sheet
                         a: append mode, write the new date below the old data
                    xls_nm: the file path of excel
        '''
        mode = kwargs.get('mode', 'w')
        xls_nm = kwargs.get('xls_nm', self.xls_nm)

        if os.path.isfile(xls_nm):
            wb = load_workbook(xls_nm)
        else:
            wb = Workbook()
        if sheet_name in wb.sheetnames and mode == 'a':
            ws = wb[sheet_name]
        else:
            if sheet_name in wb.sheetnames and mode == 'w':
                wb.remove(wb[sheet_name])
            ws = wb.create_sheet(sheet_name, 0)
        for l in list_data:
            try:
                ws.append(l)
            except:
                self.logger.error(l + ' can not append in Sheet ' + sheet_name)
        wb.save(xls_nm)


    def read_sheet(self, sheet_nm, data_s='list', **kwargs):
        '''
        description: read sheet to data structure in [dict,sheet,dinl]
        author: Kail
        params: sheet_nm, data_s
                data_s: data structure, choosen from [list, dict, dinl], default is 'list'
                sheet_nm: the Sheet name which want to read
                kwargs: optional keyword value pairs
                    xls_nm: the file path of excel
        return: data according to variable 'data_s'
        '''

        xls_nm = kwargs.get('xls_nm', self.xls_nm)
        max_row = kwargs.get('max_row')

        wb = load_workbook(xls_nm, data_only=True)
        ws = wb[sheet_nm]
        d_list = []
        r_num = 1
        for row in ws.rows:
            if max_row:
                if r_num > max_row:
                    break
            row_list = [cell.value for cell in row]
            tmp_str = ''.join([str(v) for v in row_list if v != None and re.search(r'\S', str(v))])
            if re.search(r'\S', tmp_str):
                d_list.append(row_list)
            else:
                break
            r_num += 1
        if data_s == 'list':
            return d_list
        elif data_s == 'dict':
            return self.csv_parser.date_trans(data=d_list, trans_type='list_dict')
        elif data_s == 'dinl':
            return self.csv_parser.date_trans(data=d_list, trans_type='list_dinl')


    def del_cloumn(self, raw_col, sheet_nm, **kwargs):
        '''
        description: delete Sheet's rows when cell value is 'uwant_colv' in column 'raw_col'
                     and saved the delete rows in new sheet
        author: Kail
        params: raw_col: the column seleted
                sheet_nm: the Sheet name which want to read
                kwargs: optional keyword value pairs
                    xls_nm: the file path of excel, default: self.xls_nm
                    uwant_colv: the cell value wanted to delete in column
                    del_row: if keep old value in sheet
                        choices: (True/False), default True
                        True: delete select rows in sheet
                        False: keep select rows in sheet
                    saved_sheet: the sheet name which saved the deleted rows
        return: unwanted list
        '''
        xls_nm = kwargs.get('xls_nm', self.xls_nm)
        uwant_colv = kwargs.get('uwant_colv', 'N')
        del_raw = kwargs.get('del_raw', True)
        saved_sheet = kwargs.get('saved_sheet')

        unwant_list = []
        data = self.read_sheet(sheet_nm=sheet_nm, data_s='list')
        unwant_list = [data[0]]
        col_idx = data[0].index(raw_col)
        for lis in data[1:]:
            if lis[col_idx] == uwant_colv:
                unwant_list.append(lis)
                if del_raw:
                    data.remove(lis)
        if saved_sheet:
            self.list_2sheet(list_data=unwant_list, xls_nm=xls_nm, sheet_name=saved_sheet)
        else:
            self.list_2sheet(list_data=unwant_list, xls_nm=xls_nm, sheet_name=raw_col + '=' + uwant_colv)
        self.list_2sheet(list_data=data, xls_nm=xls_nm, sheet_name=sheet_nm)
        return unwant_list


    def rep_column(self, raw_col, rep_col, sheet_nm, **kwargs):
        '''
        description: replace Sheet's column value from 'raw_col' to 'rep_col' when
                        the value in 'raw_col' match 'search_rgx'
        author: Kail
        params: raw_col, the check column
                rep_col: the column want to replace to
                sheet_nm: the Sheet name which want to read
                kwargs: optional keyword value pairs
                    xls_nm: the file path of excel, default: self.xls_nm
                    search_rgx: the mapping rules of values in raw cloumn
        '''
        search_rgx = kwargs.get('search_rgx', r'(80|13)\s*')
        xls_nm = kwargs.get('xls_nm', self.xls_nm)

        data = self.read_sheet(sheet_nm=sheet_nm, data_s='list')
        raw_cidx = data[0].index(raw_col)
        rep_cidx = data[0].index(rep_col)
        unnormal_rows = [data[0]]
        for lis in data[1:]:
            if lis[raw_cidx] == '?':
                continue
            elif re.match(search_rgx, lis[raw_cidx]):
                lis[rep_cidx] = lis[raw_cidx]
            else:
                unnormal_rows.append(lis)
                data.remove(lis)
        self.list_2sheet(list_data=data, xls_nm=xls_nm, sheet_name=sheet_nm)
        if len(unnormal_rows) > 1:
            unnormal_rows.insert(0, ["Below rows column '%s'" %raw_col +
                                     " is neither '?' nor 'N' and is" +
                                     " not partnumber start with (80|13) "]
            )
            self.list_2sheet(list_data=unnormal_rows, xls_nm=xls_nm, sheet_name='abnormal_rows', mode='a')


    def sheet_action(self, sheet_nm, action, **kwargs):
        '''
        description: doing [add/delete][rows/columns] action
        author: Kail
        params: action, the sheet action, choices:
                    [insert_rows, delete_rows, insert_cols, delete_cols]
                start_idx: the action start index unmber
                start_head: the [row/column] head value
                sheet_nm: the Sheet name which want to read
                kwargs: optional keyword value pairs
                    xls_nm: the file path of excel, default: self.xls_nm
                    num:    the [insert/delete] number
                    saved_excel: the excel saved name, default 'xls_nm'
        '''
        xls_nm = kwargs.get('xls_nm', self.xls_nm)
        start_idx = kwargs.get('start_idx')
        num = kwargs.get('num', 1)
        saved_excel = kwargs.get('saved_excel')
        start_head = kwargs.get('start_head')

        wb = load_workbook(xls_nm)
        ws = wb[sheet_nm]
        title_list = [cell.value for cell in ws[1]]
        if start_head:
            start_idx = title_list.index(start_head)
        if action == 'delete_rows':
            ws.delete_rows(start_idx, num)
        elif action == 'delete_cols':
            ws.delete_cols(start_idx, num)
        elif action == 'insert_rows':
            ws.insert_rows(start_idx, num)
        elif action == 'insert_cols':
            ws.insert_cols(start_idx, num)

        if saved_excel:
            wb.save(saved_excel)
        else:
            wb.save(xls_nm)


    def sheet_info(self, sheet_nm, *inf, **kwargs):
        '''
        description: get sheet infomation
        author: Kail
        params: inf, the infomation want to get, choices:
                    ['max_row', 'max_col']
                sheet_nm: the Sheet name
                kwargs: optional keyword value pairs
                    xls_nm: the file path of excel, default: self.xls_nm
        return: the infomation tuple according to '*inf'
        '''
        xls_nm = kwargs.get('xls_nm', self.xls_nm)
        wb = load_workbook(xls_nm)
        ws = wb[sheet_nm]

        max_row = ws.max_row
        max_col = ws.max_column
        ret_list = []
        for i in inf:
            ret_list.append(eval(i))
        return tuple(ret_list)


    def set_col_width(ws, width, **kwargs):
        '''
        description: set sheet column width
        author: Kail
        params: ws: worksheet
                width: the column's with
                kwargs: optional keyword value pairs
        return: None
        '''
        for i in range(1,ws.max_column+1):
            letter=get_column_letter(i)
            ws.column_dimensions[letter].width=width
